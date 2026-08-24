import io
from datetime import datetime, date, timedelta
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from projects.models import Task, TaskDependency, Project


def export_project_to_excel(project: Project) -> io.BytesIO:
    """
    Generates a stylized multi-sheet Excel (.xlsx) file representing the full
    Milestone Management project schedule, complete with:
    1. 'Gantt Chart & Timeline' - Side-by-side data table + colored calendar Gantt bars
    2. 'Task Details & Financials' - Full data sheet (INR ₹, DD/MM/YYYY, Man-Hours, EVM)
    3. 'Executive KPI Summary' - Project health, EVM metrics, and performance indexes
    """
    wb = openpyxl.Workbook()
    
    # --------------------------------------------------------------------------
    # Styles & Colors Palette
    # --------------------------------------------------------------------------
    title_font = Font(name="Segoe UI", size=15, bold=True, color="0F172A")
    subtitle_font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    section_font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    day_header_font = Font(name="Segoe UI", size=8, bold=True, color="475569")
    month_header_font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
    regular_font = Font(name="Segoe UI", size=9, color="334155")
    small_font = Font(name="Segoe UI", size=8, color="64748B")
    bar_text_font = Font(name="Segoe UI", size=8, bold=True, color="FFFFFF")
    milestone_font = Font(name="Segoe UI", size=11, bold=True, color="D97706")

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    month_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    day_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    weekend_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    parent_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Gantt Bar fills
    bar_normal = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")     # Blue
    bar_complete = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")   # Emerald Green
    bar_critical = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid")   # Rose Red
    bar_parent = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")     # Dark Slate
    bar_milestone = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber Light

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    timeline_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    tasks = list(project.tasks.select_related('assignee', 'parent_task').prefetch_related('predecessors').order_by('sort_order', 'id'))

    # ==========================================================================
    # SHEET 1: Gantt Chart & Visual Timeline
    # ==========================================================================
    ws_gantt = wb.active
    ws_gantt.title = "Gantt Chart & Schedule"
    ws_gantt.views.sheetView[0].showGridLines = True

    # Title Block
    ws_gantt['A1'] = f"MILESTONE MANAGEMENT — {project.name.upper()}"
    ws_gantt['A1'].font = title_font
    ws_gantt['A2'] = f"Project Code: {project.code} | Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')} IST | Currency: INR (₹) | Standard: Indian Business Calendar"
    ws_gantt['A2'].font = subtitle_font

    # Determine timeline date range
    all_start_dates = [t.start_date for t in tasks if t.start_date]
    all_end_dates = [t.end_date for t in tasks if t.end_date]
    today = date.today()

    if all_start_dates:
        min_date = min(all_start_dates)
    else:
        min_date = today

    if all_end_dates:
        max_date = max(all_end_dates)
    else:
        max_date = min_date + timedelta(days=30)

    # Pad timeline range by 2 days before and 5 days after for breathing room
    min_date = min_date - timedelta(days=2)
    max_date = max_date + timedelta(days=5)
    total_days = max(14, min(120, (max_date - min_date).days + 1))  # Cap between 14 to 120 days for optimal Excel layout

    timeline_dates = [min_date + timedelta(days=i) for i in range(total_days)]

    # Left task columns headers
    gantt_headers = [
        "WBS", "Task Name", "Assignee", "Start Date", "End Date",
        "Days", "Progress", "Status", "Critical", "Milestone"
    ]
    num_task_cols = len(gantt_headers)

    # Legend in Row 1 & 2 (Right side)
    legend_start_col = num_task_cols + 2
    legends = [
        ("In Progress", "3B82F6", "FFFFFF"),
        ("Completed (100%)", "10B981", "FFFFFF"),
        ("Critical Path", "E11D48", "FFFFFF"),
        ("Milestone (◆)", "FEF3C7", "D97706"),
        ("Summary Group", "1E293B", "FFFFFF")
    ]
    for idx, (leg_name, bg_hex, fg_hex) in enumerate(legends):
        col_c = legend_start_col + (idx * 3)
        if col_c + 2 <= num_task_cols + total_days + 1:
            ws_gantt.merge_cells(start_row=2, start_column=col_c, end_row=2, end_column=col_c + 2)
            c = ws_gantt.cell(row=2, column=col_c, value=leg_name)
            c.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
            c.font = Font(name="Segoe UI", size=8, bold=True, color=fg_hex)
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Write Column Headers in Row 4
    for col_idx, h in enumerate(gantt_headers, 1):
        ws_gantt.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        cell = ws_gantt.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Build Month / Year Headers (Row 3) and Day Headers (Row 4) for Timeline
    month_spans = []
    current_month_key = None
    span_start = None

    for day_idx, d in enumerate(timeline_dates):
        timeline_col = num_task_cols + 1 + day_idx
        m_key = d.strftime('%B %Y')
        if m_key != current_month_key:
            if current_month_key is not None:
                month_spans.append((current_month_key, span_start, timeline_col - 1))
            current_month_key = m_key
            span_start = timeline_col
        
        # Day Header in Row 4
        d_cell = ws_gantt.cell(row=4, column=timeline_col, value=f"{d.day:02d}\n{d.strftime('%a')[:2]}")
        d_cell.font = day_header_font
        is_weekend = d.weekday() >= 5
        d_cell.fill = PatternFill(start_color="E2E8F0" if is_weekend else "F1F5F9", fill_type="solid")
        d_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        d_cell.border = timeline_border
        
        # Set narrow column width for timeline day cells
        col_letter = get_column_letter(timeline_col)
        ws_gantt.column_dimensions[col_letter].width = 4.2

    if current_month_key is not None:
        month_spans.append((current_month_key, span_start, num_task_cols + total_days))

    # Apply Merged Month Headers (Row 3)
    for m_label, start_c, end_c in month_spans:
        if start_c <= end_c:
            ws_gantt.merge_cells(start_row=3, start_column=start_c, end_row=3, end_column=end_c)
            m_cell = ws_gantt.cell(row=3, column=start_c, value=m_label.upper())
            m_cell.font = month_header_font
            m_cell.fill = month_fill
            m_cell.alignment = Alignment(horizontal="center", vertical="center")
            m_cell.border = timeline_border

    # Write Tasks & Gantt Timeline Bars
    start_row = 5
    for row_offset, task in enumerate(tasks):
        current_row = start_row + row_offset
        ws_gantt.row_dimensions[current_row].height = 20

        # WBS indentation
        indent_level = task.wbs_code.count('.')
        name_display = ("    " * indent_level) + task.name
        is_parent = task.is_parent

        # Task Data Columns
        ws_gantt.cell(row=current_row, column=1, value=task.wbs_code).alignment = Alignment(horizontal="center", vertical="center")
        ws_gantt.cell(row=current_row, column=2, value=name_display).alignment = Alignment(horizontal="left", vertical="center")
        ws_gantt.cell(row=current_row, column=3, value=task.assignee.get_full_name() or task.assignee.username if task.assignee else "-").alignment = Alignment(horizontal="left", vertical="center")
        ws_gantt.cell(row=current_row, column=4, value=task.start_date.strftime('%d/%m/%Y') if task.start_date else "").alignment = Alignment(horizontal="center", vertical="center")
        ws_gantt.cell(row=current_row, column=5, value=task.end_date.strftime('%d/%m/%Y') if task.end_date else "").alignment = Alignment(horizontal="center", vertical="center")
        ws_gantt.cell(row=current_row, column=6, value=task.duration_days).alignment = Alignment(horizontal="center", vertical="center")
        ws_gantt.cell(row=current_row, column=7, value=f"{task.progress}%").alignment = Alignment(horizontal="center", vertical="center")
        ws_gantt.cell(row=current_row, column=8, value=task.get_status_display()).alignment = Alignment(horizontal="center", vertical="center")
        ws_gantt.cell(row=current_row, column=9, value="CRITICAL" if task.is_critical else "Normal").alignment = Alignment(horizontal="center", vertical="center")
        ws_gantt.cell(row=current_row, column=10, value="◆ Milestone" if task.is_milestone else "Task").alignment = Alignment(horizontal="center", vertical="center")

        # Style Task Data Columns
        for col_idx in range(1, num_task_cols + 1):
            c = ws_gantt.cell(row=current_row, column=col_idx)
            c.border = thin_border
            c.font = bold_font if is_parent else regular_font
            if is_parent:
                c.fill = parent_fill
            elif task.is_critical and col_idx in [1, 2, 9]:
                c.fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")

        # Fill Visual Gantt Timeline Cells for this task
        t_start = task.start_date
        t_end = task.end_date or t_start

        for day_idx, d in enumerate(timeline_dates):
            timeline_col = num_task_cols + 1 + day_idx
            t_cell = ws_gantt.cell(row=current_row, column=timeline_col)
            t_cell.border = timeline_border
            is_weekend = d.weekday() >= 5

            if t_start and t_end and (t_start <= d <= t_end):
                if task.is_milestone:
                    t_cell.value = "◆"
                    t_cell.font = milestone_font
                    t_cell.fill = bar_milestone
                    t_cell.alignment = Alignment(horizontal="center", vertical="center")
                elif is_parent:
                    t_cell.fill = bar_parent
                elif task.is_critical:
                    t_cell.fill = bar_critical
                    if d == t_start:
                        t_cell.value = f"{task.progress}%"
                        t_cell.font = bar_text_font
                        t_cell.alignment = Alignment(horizontal="center", vertical="center")
                elif task.progress == 100 or task.status == 'COMPLETE':
                    t_cell.fill = bar_complete
                    if d == t_start:
                        t_cell.value = "✓"
                        t_cell.font = bar_text_font
                        t_cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    t_cell.fill = bar_normal
                    if d == t_start:
                        t_cell.value = f"{task.progress}%"
                        t_cell.font = bar_text_font
                        t_cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if is_weekend:
                    t_cell.fill = weekend_fill

    # Set column widths for Task Data Columns
    data_widths = [8, 30, 16, 13, 13, 8, 10, 14, 11, 12]
    for idx, width in enumerate(data_widths, 1):
        col_letter = get_column_letter(idx)
        ws_gantt.column_dimensions[col_letter].width = width

    # ==========================================================================
    # SHEET 2: Task Details & Financials (Full Data)
    # ==========================================================================
    ws_data = wb.create_sheet(title="Task Details & Financials")
    ws_data.views.sheetView[0].showGridLines = True

    ws_data['A1'] = f"MILESTONE MANAGEMENT — DETAILED TASK SHEET"
    ws_data['A1'].font = title_font
    ws_data['A2'] = f"Project: {project.name} ({project.code}) | Currency: INR (₹) | EVM Baseline Enabled"
    ws_data['A2'].font = subtitle_font

    full_headers = [
        "WBS", "Task Name", "Hierarchy Level", "Assignee", "Start Date", "End Date",
        "Duration (Days)", "Progress %", "Status", "Priority", "Milestone",
        "Critical Path", "Predecessors", "Baseline Start", "Baseline End",
        "Schedule Variance (Days)", "Estimated Cost (₹ INR)", "Actual Cost (₹ INR)",
        "Est. Effort (Hours)", "Actual Effort (Hours)"
    ]

    header_row = 4
    for col_num, header_title in enumerate(full_headers, 1):
        cell = ws_data.cell(row=header_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    current_row = 5
    for task in tasks:
        preds = task.predecessors.values_list('from_task__name', flat=True)
        preds_str = ", ".join(preds)

        ws_data.cell(row=current_row, column=1, value=task.wbs_code)
        ws_data.cell(row=current_row, column=2, value=task.name)
        ws_data.cell(row=current_row, column=3, value=task.wbs_code.count('.'))
        ws_data.cell(row=current_row, column=4, value=task.assignee.get_full_name() or task.assignee.username if task.assignee else "Unassigned")
        ws_data.cell(row=current_row, column=5, value=task.start_date.strftime('%d/%m/%Y') if task.start_date else "")
        ws_data.cell(row=current_row, column=6, value=task.end_date.strftime('%d/%m/%Y') if task.end_date else "")
        ws_data.cell(row=current_row, column=7, value=task.duration_days)
        ws_data.cell(row=current_row, column=8, value=f"{task.progress}%")
        ws_data.cell(row=current_row, column=9, value=task.get_status_display())
        ws_data.cell(row=current_row, column=10, value=task.get_priority_display())
        ws_data.cell(row=current_row, column=11, value="Yes" if task.is_milestone else "No")
        ws_data.cell(row=current_row, column=12, value="YES" if task.is_critical else "No")
        ws_data.cell(row=current_row, column=13, value=preds_str)
        ws_data.cell(row=current_row, column=14, value=task.baseline_start_date.strftime('%d/%m/%Y') if task.baseline_start_date else "-")
        ws_data.cell(row=current_row, column=15, value=task.baseline_end_date.strftime('%d/%m/%Y') if task.baseline_end_date else "-")
        ws_data.cell(row=current_row, column=16, value=task.schedule_variance_days)

        c_est = ws_data.cell(row=current_row, column=17, value=float(task.estimated_cost))
        c_est.number_format = '₹#,##,##0.00'
        c_act = ws_data.cell(row=current_row, column=18, value=float(task.actual_cost))
        c_act.number_format = '₹#,##,##0.00'

        ws_data.cell(row=current_row, column=19, value=task.estimated_hours)
        ws_data.cell(row=current_row, column=20, value=task.actual_hours)

        is_parent = task.is_parent
        for col_num in range(1, len(full_headers) + 1):
            c = ws_data.cell(row=current_row, column=col_num)
            c.border = thin_border
            c.font = bold_font if is_parent else regular_font
            if task.is_critical and col_num in [1, 2, 12]:
                c.fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
            elif is_parent:
                c.fill = parent_fill

        current_row += 1

    # Auto-adjust column widths for Sheet 2
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ==========================================================================
    # SHEET 3: Executive KPI Summary & EVM Dashboard
    # ==========================================================================
    ws_kpi = wb.create_sheet(title="Executive KPI Summary")
    ws_kpi.views.sheetView[0].showGridLines = True

    ws_kpi['A1'] = f"PROJECT PERFORMANCE & EVM DASHBOARD — {project.name.upper()}"
    ws_kpi['A1'].font = title_font
    ws_kpi['A2'] = f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')} IST | Master Status Report"
    ws_kpi['A2'].font = subtitle_font

    # Summary Metrics Calculations
    total_tasks = len(tasks)
    completed_tasks = sum(1 for t in tasks if t.status == 'COMPLETE' or t.progress == 100)
    critical_tasks = sum(1 for t in tasks if t.is_critical)
    milestones_count = sum(1 for t in tasks if t.is_milestone)
    
    total_pv = sum(t.estimated_cost for t in tasks)
    total_ev = sum(t.estimated_cost * Decimal(t.progress) / Decimal(100) for t in tasks)
    total_ac = sum(t.actual_cost for t in tasks)

    cpi = float(total_ev / total_ac) if total_ac > 0 else 1.0
    spi = float(total_ev / total_pv) if total_pv > 0 else 1.0
    cost_variance = total_ev - total_ac
    schedule_variance = total_ev - total_pv

    kpi_cards = [
        ("Project Code", project.code, "General Info"),
        ("Project Status", project.get_status_display(), "Health"),
        ("Total Tasks / Milestones", f"{total_tasks} ({milestones_count} Milestones)", "Scope"),
        ("Completed Tasks", f"{completed_tasks} ({round(completed_tasks/total_tasks*100) if total_tasks else 0}%)", "Progress"),
        ("Critical Path Tasks", f"{critical_tasks} tasks", "Schedule Risk"),
        ("Planned Value (PV - Budget)", f"₹ {total_pv:,.2f}", "EVM Financials"),
        ("Earned Value (EV - Realized)", f"₹ {total_ev:,.2f}", "EVM Financials"),
        ("Actual Cost (AC - Spent)", f"₹ {total_ac:,.2f}", "EVM Financials"),
        ("Cost Variance (CV = EV - AC)", f"₹ {cost_variance:,.2f}", "EVM Financials"),
        ("Schedule Variance (SV = EV - PV)", f"₹ {schedule_variance:,.2f}", "EVM Financials"),
        ("Cost Performance Index (CPI)", f"{cpi:.2f} ({'Under Budget' if cpi >= 1 else 'Over Budget'})", "Efficiency"),
        ("Schedule Performance Index (SPI)", f"{spi:.2f} ({'Ahead/On Track' if spi >= 1 else 'Behind Schedule'})", "Efficiency"),
    ]

    ws_kpi['A4'] = "METRIC NAME"
    ws_kpi['B4'] = "VALUE"
    ws_kpi['C4'] = "CATEGORY / ASSESSMENT"
    for col_idx, h_text in enumerate(["METRIC NAME", "VALUE", "CATEGORY / ASSESSMENT"], 1):
        cell = ws_kpi.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if col_idx != 2 else "center", vertical="center")
        cell.border = thin_border

    kpi_row = 5
    for name_val, val_text, cat_text in kpi_cards:
        c1 = ws_kpi.cell(row=kpi_row, column=1, value=name_val)
        c2 = ws_kpi.cell(row=kpi_row, column=2, value=str(val_text))
        c3 = ws_kpi.cell(row=kpi_row, column=3, value=cat_text)

        c1.font = bold_font
        c2.font = bold_font if "CPI" in name_val or "SPI" in name_val or "Cost" in name_val else regular_font
        c3.font = small_font

        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border

        if "CPI" in name_val or "SPI" in name_val:
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.fill = PatternFill(start_color="ECFDF5" if ("Under" in str(val_text) or "Ahead" in str(val_text)) else "FEF2F2", fill_type="solid")

        kpi_row += 1

    ws_kpi.column_dimensions['A'].width = 36
    ws_kpi.column_dimensions['B'].width = 30
    ws_kpi.column_dimensions['C'].width = 28

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output



def import_project_from_excel(file_obj, project: Project) -> int:
    """
    Imports or updates tasks from an uploaded Excel (.xlsx) file with support for
    both DD/MM/YYYY (Indian Standard) and standard ISO date formats.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=col_idx).value
        if val:
            headers[str(val).strip().lower()] = col_idx

    task_name_col = None
    for h, idx in headers.items():
        if "task name" in h or "name" in h:
            task_name_col = idx
            break

    start_col = None
    for h, idx in headers.items():
        if "start" in h and "baseline" not in h:
            start_col = idx
            break

    end_col = None
    for h, idx in headers.items():
        if "end" in h and "baseline" not in h:
            end_col = idx
            break

    progress_col = None
    for h, idx in headers.items():
        if "progress" in h:
            progress_col = idx
            break

    imported_count = 0
    today = date.today()

    for row_idx in range(5, ws.max_row + 1):
        name_val = ws.cell(row=row_idx, column=task_name_col or 2).value if task_name_col else ws.cell(row=row_idx, column=2).value
        if not name_val:
            continue

        name_str = str(name_val).strip()
        
        # Parse start date
        start_val = ws.cell(row=row_idx, column=start_col or 5).value if start_col else None
        if isinstance(start_val, (datetime, date)):
            start_d = start_val.date() if isinstance(start_val, datetime) else start_val
        elif isinstance(start_val, str):
            try:
                # Try DD/MM/YYYY (Indian Standard)
                start_d = datetime.strptime(start_val[:10], '%d/%m/%Y').date()
            except ValueError:
                try:
                    start_d = datetime.strptime(start_val[:10], '%Y-%m-%d').date()
                except ValueError:
                    start_d = today
        else:
            start_d = today

        # Parse end date
        end_val = ws.cell(row=row_idx, column=end_col or 6).value if end_col else None
        if isinstance(end_val, (datetime, date)):
            end_d = end_val.date() if isinstance(end_val, datetime) else end_val
        elif isinstance(end_val, str):
            try:
                # Try DD/MM/YYYY (Indian Standard)
                end_d = datetime.strptime(end_val[:10], '%d/%m/%Y').date()
            except ValueError:
                try:
                    end_d = datetime.strptime(end_val[:10], '%Y-%m-%d').date()
                except ValueError:
                    end_d = start_d
        else:
            end_d = start_d

        # Parse progress
        prog_val = ws.cell(row=row_idx, column=progress_col or 8).value if progress_col else 0
        try:
            prog_int = int(str(prog_val).replace('%', '').strip())
        except (ValueError, TypeError):
            prog_int = 0

        Task.objects.create(
            project=project,
            name=name_str,
            start_date=start_d,
            end_date=end_d,
            progress=max(0, min(100, prog_int)),
            sort_order=imported_count + 1
        )
        imported_count += 1

    return imported_count
